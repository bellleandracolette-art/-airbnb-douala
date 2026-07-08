import json
import random
from pathlib import Path
from datetime import datetime
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

random.seed(42)

OUTPUT = Path.home() / "Desktop" / "airbnb-douala" / "data" / "processed" / "Airbnb_Douala_Details.xlsx"

conn = mysql.connector.connect(host="localhost", user="root", password="root", database="airbnb_douala")
cursor = conn.cursor(dictionary=True)

cursor.execute("SELECT * FROM listings ORDER BY source, neighborhood")
listings = cursor.fetchall()

cursor.execute("SELECT * FROM stats_quartiers")
stats_q = cursor.fetchall()

cursor.execute("SELECT * FROM stats_types")
stats_t = cursor.fetchall()

conn.close()

print("Loaded " + str(len(listings)) + " listings from MySQL")

# ── Enrichir les donnees avec des attributs Airbnb simules realistes ─────────
def enrich(row):
    price = row.get("price_per_night") or 20000
    typ = (row.get("type") or "").lower()
    hood = (row.get("neighborhood") or "Douala").title()
    source = row.get("source") or "simulated"

    # Surface m2 estimee selon type et prix
    if "villa" in typ:
        surface = random.randint(120, 350)
        chambres = random.randint(3, 6)
        lits = chambres + random.randint(0, 2)
        sdb = random.randint(2, 4)
        capacite = chambres * 2
    elif "appartement entier" in typ or "appartement" in typ:
        surface = random.randint(35, 120)
        chambres = random.randint(1, 3)
        lits = chambres + random.randint(0, 1)
        sdb = max(1, chambres - 1)
        capacite = chambres * 2
    elif "studio" in typ:
        surface = random.randint(18, 40)
        chambres = 0
        lits = random.randint(1, 2)
        sdb = 1
        capacite = random.randint(1, 2)
    elif "chambre" in typ:
        surface = random.randint(12, 30)
        chambres = 1
        lits = 1
        sdb = 1
        capacite = random.randint(1, 2)
    else:
        surface = random.randint(30, 100)
        chambres = random.randint(1, 3)
        lits = chambres
        sdb = 1
        capacite = chambres * 2

    # Equipements selon prix et quartier premium
    premium = hood in ["Bonanjo", "Bonapriso", "Bali"]
    wifi = True  # quasi universel
    parking_prob = 0.75 if premium else 0.45
    parking_gratuit = random.random() < parking_prob
    parking = "Gratuit" if parking_gratuit else ("Payant" if random.random() < 0.2 else "Non")
    piscine = random.random() < (0.35 if "villa" in typ else 0.05)
    clim = random.random() < (0.95 if premium else 0.7)
    cuisine = random.random() < (0.9 if "entier" in typ or "villa" in typ or "studio" in typ else 0.3)
    lave_linge = random.random() < (0.7 if "entier" in typ or "villa" in typ else 0.2)
    tv = random.random() < 0.8
    generateur = random.random() < (0.6 if premium else 0.35)
    securite = random.random() < (0.7 if premium else 0.4)
    ascenseur = random.random() < (0.3 if "appartement" in typ else 0.05)
    balcon = random.random() < (0.5 if premium else 0.25)
    jardin = random.random() < (0.4 if "villa" in typ or "maison" in typ else 0.05)
    animaux = random.random() < 0.15
    fumeurs = random.random() < 0.1
    annulation_gratuite = random.random() < 0.65
    sejour_min = random.choice([1, 1, 1, 2, 2, 3, 5, 7])
    superhost = bool(row.get("superhost")) or random.random() < 0.25
    note = row.get("rating") or round(random.uniform(3.8, 5.0), 2)
    avis = row.get("reviews") or random.randint(2, 120)
    occ = row.get("occupancy_rate") or round(random.uniform(0.45, 0.85), 2)
    revenu = row.get("monthly_revenue") or round(price * 30 * occ)

    return {
        "id": row.get("id"),
        "source": source,
        "titre": (row.get("title") or "")[:80],
        "quartier": hood,
        "type_bien": row.get("type") or "Appartement",
        "surface_m2": surface,
        "nb_chambres": chambres,
        "nb_lits": lits,
        "nb_salles_bain": sdb,
        "capacite_max": capacite,
        "prix_nuit_xaf": int(price),
        "prix_nuit_eur": round(price / 655.957, 0),
        "note": note,
        "nb_avis": avis,
        "superhost": "Oui" if superhost else "Non",
        "taux_occupation": str(round(occ * 100)) + "%",
        "revenu_mensuel_xaf": int(revenu),
        "sejour_min_nuits": sejour_min,
        "annulation_gratuite": "Oui" if annulation_gratuite else "Non",
        "wifi": "Oui" if wifi else "Non",
        "parking": parking,
        "piscine": "Oui" if piscine else "Non",
        "climatisation": "Oui" if clim else "Non",
        "cuisine_equipee": "Oui" if cuisine else "Non",
        "lave_linge": "Oui" if lave_linge else "Non",
        "television": "Oui" if tv else "Non",
        "generateur": "Oui" if generateur else "Non",
        "securite_gardien": "Oui" if securite else "Non",
        "ascenseur": "Oui" if ascenseur else "Non",
        "balcon_terrasse": "Oui" if balcon else "Non",
        "jardin": "Oui" if jardin else "Non",
        "animaux_acceptes": "Oui" if animaux else "Non",
        "fumeurs_acceptes": "Oui" if fumeurs else "Non",
    }

enriched = [enrich(r) for r in listings]
print("Enriched " + str(len(enriched)) + " listings")

# ── EXCEL ─────────────────────────────────────────────────────────────────────
wb = Workbook()

BLUE_DARK  = "1F4E79"
BLUE_MED   = "2E75B6"
BLUE_LIGHT = "D6E4F0"
WHITE      = "FFFFFF"
GRAY_ALT   = "F5F9FF"
GRAY_MED   = "DDDDDD"
GREEN_DARK = "1B5E20"
GREEN_LIGHT= "E8F5E9"
GOLD       = "F57F17"

thin = Side(style="thin", color=GRAY_MED)
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(size=10, color=WHITE, bold=True):
    return Font(name="Arial", size=size, bold=bold, color=color)

def cel(size=9, bold=False, color="111111"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(c):
    return PatternFill("solid", fgColor=c)

def ctr():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def lft():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ══ SHEET 1 — TOUTES LES ANNONCES DETAILS ════════════════════════════════════
ws1 = wb.active
ws1.title = "Annonces detaillees"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

# Title
ws1.merge_cells("A1:AF1")
ws1["A1"] = "TOUTES LES ANNONCES AIRBNB DOUALA — DETAILS COMPLETS (" + str(len(enriched)) + " logements)"
ws1["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws1["A1"].fill = fill(BLUE_DARK)
ws1["A1"].alignment = ctr()
ws1.row_dimensions[1].height = 30

# Headers
columns = [
    ("ID", 10), ("Source", 10), ("Titre", 35), ("Quartier", 14),
    ("Type de bien", 18), ("Surface m2", 11), ("Chambres", 10),
    ("Lits", 8), ("Salles de bain", 13), ("Capacite max", 12),
    ("Prix/nuit XAF", 14), ("Prix/nuit EUR", 13), ("Note", 8),
    ("Nb avis", 9), ("Superhost", 10), ("Occupation", 10),
    ("Revenu/mois XAF", 16), ("Sejour min (nuits)", 16),
    ("Annulation gratuite", 18), ("WiFi", 7), ("Parking", 10),
    ("Piscine", 9), ("Climatisation", 13), ("Cuisine equipee", 15),
    ("Lave-linge", 11), ("Television", 11), ("Generateur", 11),
    ("Securite/Gardien", 15), ("Ascenseur", 10), ("Balcon/Terrasse", 14),
    ("Jardin", 8), ("Animaux acceptes", 15), ("Fumeurs acceptes", 15),
]

keys = list(enriched[0].keys())

for col, (label, width) in enumerate(columns, 1):
    c = ws1.cell(row=2, column=col, value=label)
    c.font = hdr(9)
    c.fill = fill(BLUE_MED)
    c.alignment = ctr()
    c.border = brd
    ws1.column_dimensions[get_column_letter(col)].width = width
ws1.row_dimensions[2].height = 30

# Data rows
oui_fill = PatternFill("solid", fgColor=GREEN_LIGHT)
non_fill = PatternFill("solid", fgColor="FFF9C4")

for row_i, listing in enumerate(enriched):
    r = row_i + 3
    alt = row_i % 2 == 0
    base_fill = fill(GRAY_ALT if alt else WHITE)

    for col, key in enumerate(keys, 1):
        val = listing[key]
        c = ws1.cell(row=r, column=col, value=val)
        c.font = cel(9)
        c.border = brd

        if key == "source":
            if val == "real":
                c.font = Font(name="Arial", size=9, bold=True, color=GREEN_DARK)
                c.fill = fill(GREEN_LIGHT)
            else:
                c.fill = base_fill
            c.alignment = ctr()
        elif isinstance(val, str) and val in ["Oui", "Non", "Gratuit", "Payant"]:
            if val in ["Oui", "Gratuit"]:
                c.fill = oui_fill
                c.font = Font(name="Arial", size=9, color=GREEN_DARK)
            elif val in ["Non"]:
                c.fill = non_fill
            else:
                c.fill = base_fill
            c.alignment = ctr()
        elif col in [2, 3, 4, 5]:
            c.fill = base_fill
            c.alignment = lft()
        else:
            c.fill = base_fill
            c.alignment = ctr()

ws1.auto_filter.ref = "A2:" + get_column_letter(len(columns)) + str(len(enriched) + 2)

# ══ SHEET 2 — KPI DASHBOARD ══════════════════════════════════════════════════
ws2 = wb.create_sheet("KPI Dashboard")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:H1")
ws2["A1"] = "TABLEAU DE BORD KPI — MARCHE AIRBNB DOUALA"
ws2["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
ws2["A1"].fill = fill(BLUE_DARK)
ws2["A1"].alignment = ctr()
ws2.row_dimensions[1].height = 32

real = [e for e in enriched if e["source"] == "real"]
sim  = [e for e in enriched if e["source"] == "simulated"]

def avg(lst, key):
    vals = [v[key] for v in lst if isinstance(v[key], (int, float))]
    return round(sum(vals)/len(vals), 1) if vals else 0

kpis = [
    ("Total annonces", str(len(enriched)), BLUE_MED),
    ("Annonces reelles", str(len(real)), "2E7D32"),
    ("Prix moyen / nuit", str(avg(enriched,"prix_nuit_xaf")) + " XAF", GOLD),
    ("Note moyenne", str(avg(enriched,"note")), "6A1A9A"),
    ("Avec WiFi", str(sum(1 for e in enriched if e["wifi"]=="Oui")), "00838F"),
    ("Avec Parking gratuit", str(sum(1 for e in enriched if e["parking"]=="Gratuit")), "1565C0"),
    ("Avec Piscine", str(sum(1 for e in enriched if e["piscine"]=="Oui")), "0277BD"),
    ("Superhosts", str(sum(1 for e in enriched if e["superhost"]=="Oui")), "AD1457"),
]

for i, (label, val, color) in enumerate(kpis):
    col = (i % 4) * 2 + 1
    row = 3 + (i // 4) * 4
    c1 = get_column_letter(col)
    c2 = get_column_letter(col+1)
    ws2.merge_cells(f"{c1}{row}:{c2}{row}")
    ws2[f"{c1}{row}"] = label
    ws2[f"{c1}{row}"].font = Font(name="Arial", size=9, color="888888")
    ws2[f"{c1}{row}"].fill = fill("F0F4FF")
    ws2[f"{c1}{row}"].alignment = ctr()
    ws2.merge_cells(f"{c1}{row+1}:{c2}{row+2}")
    ws2[f"{c1}{row+1}"] = val
    ws2[f"{c1}{row+1}"].font = Font(name="Arial", size=20, bold=True, color=color)
    ws2[f"{c1}{row+1}"].fill = fill("F0F4FF")
    ws2[f"{c1}{row+1}"].alignment = ctr()

# Stats quartiers
r = 14
ws2.merge_cells("A" + str(r) + ":H" + str(r))
ws2["A" + str(r)] = "STATISTIQUES PAR QUARTIER"
ws2["A" + str(r)].font = Font(name="Arial", size=11, bold=True, color=WHITE)
ws2["A" + str(r)].fill = fill(BLUE_MED)
ws2["A" + str(r)].alignment = ctr()

hdrs_q = ["Quartier","Total","Prix moy. XAF","Note moy.","WiFi %","Parking grat. %","Piscine %","Superhost %"]
for ci, h in enumerate(hdrs_q, 1):
    c = ws2.cell(row=r+1, column=ci, value=h)
    c.font = hdr(9)
    c.fill = fill(BLUE_DARK)
    c.alignment = ctr()
    c.border = brd

from collections import defaultdict
by_hood = defaultdict(list)
for e in enriched:
    by_hood[e["quartier"]].append(e)

for ri, (hood, items) in enumerate(sorted(by_hood.items(), key=lambda x: -len(x[1]))):
    row_r = r + 2 + ri
    alt = ri % 2 == 0
    bf = fill(GRAY_ALT if alt else WHITE)
    vals = [
        hood, len(items),
        avg(items, "prix_nuit_xaf"),
        avg(items, "note"),
        str(round(sum(1 for e in items if e["wifi"]=="Oui")/len(items)*100)) + "%",
        str(round(sum(1 for e in items if e["parking"]=="Gratuit")/len(items)*100)) + "%",
        str(round(sum(1 for e in items if e["piscine"]=="Oui")/len(items)*100)) + "%",
        str(round(sum(1 for e in items if e["superhost"]=="Oui")/len(items)*100)) + "%",
    ]
    for ci, val in enumerate(vals, 1):
        c = ws2.cell(row=row_r, column=ci, value=val)
        c.font = cel(9)
        c.fill = bf
        c.alignment = ctr()
        c.border = brd

set_widths(ws2, {"A":18,"B":10,"C":16,"D":10,"E":10,"F":15,"G":12,"H":13})

# ══ SHEET 3 — STATS QUARTIERS ════════════════════════════════════════════════
ws3 = wb.create_sheet("Stats quartiers")
ws3.sheet_view.showGridLines = False
ws3.merge_cells("A1:G1")
ws3["A1"] = "STATISTIQUES PAR QUARTIER — MySQL"
ws3["A1"].font = Font(name="Arial", size=12, bold=True, color=WHITE)
ws3["A1"].fill = fill(BLUE_DARK)
ws3["A1"].alignment = ctr()

hdrs3 = ["Quartier","Total annonces","Prix min XAF","Prix moyen XAF","Prix max XAF","Note moyenne","Annonces reelles"]
for ci, h in enumerate(hdrs3, 1):
    c = ws3.cell(row=2, column=ci, value=h)
    c.font = hdr(10)
    c.fill = fill(BLUE_MED)
    c.alignment = ctr()
    c.border = brd

for ri, row in enumerate(stats_q):
    r = ri + 3
    alt = ri % 2 == 0
    vals = [row["neighborhood"], row["total_annonces"], row["prix_min"], row["prix_moyen"], row["prix_max"], row["note_moyenne"], row["annonces_reelles"]]
    for ci, val in enumerate(vals, 1):
        c = ws3.cell(row=r, column=ci, value=val)
        c.font = cel(10)
        c.fill = fill(GRAY_ALT if alt else WHITE)
        c.alignment = ctr()
        c.border = brd

set_widths(ws3, {"A":16,"B":15,"C":14,"D":16,"E":14,"F":13,"G":16})

# Chart
chart = BarChart()
chart.type = "col"
chart.title = "Prix moyen par quartier (XAF)"
chart.y_axis.title = "Prix XAF"
chart.style = 10
data = Reference(ws3, min_col=4, min_row=2, max_row=2+len(stats_q))
cats = Reference(ws3, min_col=1, min_row=3, max_row=2+len(stats_q))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 20
chart.height = 12
ws3.add_chart(chart, "I2")

wb.save(str(OUTPUT))
print("")
print("Excel saved: " + str(OUTPUT))
print("Sheet 1: " + str(len(enriched)) + " annonces avec 33 colonnes de details")
print("Sheet 2: KPI Dashboard")
print("Sheet 3: Stats quartiers + graphique")

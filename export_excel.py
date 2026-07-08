import sys
from pathlib import Path
sys.path.insert(0, str(Path.home() / "Desktop" / "airbnb-douala"))

import pandas as pd
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

OUTPUT = Path.home() / "Desktop" / "airbnb-douala" / "data" / "processed" / "Analyse_Airbnb_Douala.xlsx"

conn = mysql.connector.connect(host="localhost", user="root", password="root", database="airbnb_douala")

df_all = pd.read_sql("SELECT * FROM listings", conn)
df_quartiers = pd.read_sql("SELECT * FROM stats_quartiers", conn)
df_types = pd.read_sql("SELECT * FROM stats_types", conn)
df_real = pd.read_sql("SELECT id, title, neighborhood, type, price_per_night, rating, reviews, source FROM listings WHERE source='real' ORDER BY price_per_night DESC", conn)
conn.close()

wb = Workbook()

# ── COLORS ────────────────────────────────────────────────────────────────────
BLUE_DARK  = "1F4E79"
BLUE_MED   = "2E75B6"
BLUE_LIGHT = "D6E4F0"
GOLD       = "FFB400"
GREEN      = "2E7D32"
WHITE      = "FFFFFF"
GRAY_LIGHT = "F5F9FF"
GRAY_MED   = "DDDDDD"

def hdr_font(size=11, color=WHITE, bold=True):
    return Font(name="Arial", size=size, bold=bold, color=color)

def cell_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(color):
    return PatternFill("solid", fgColor=color)

def border():
    thin = Side(style="thin", color=GRAY_MED)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def style_header_row(ws, row, cols, bg=BLUE_MED):
    for col in cols:
        c = ws.cell(row=row, column=col)
        c.font = hdr_font()
        c.fill = fill(bg)
        c.alignment = center()
        c.border = border()

def style_data_row(ws, row, cols, alt=False):
    bg = GRAY_LIGHT if alt else WHITE
    for col in cols:
        c = ws.cell(row=row, column=col)
        c.font = cell_font()
        c.fill = fill(bg)
        c.alignment = center()
        c.border = border()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — TABLEAU DE BORD
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Tableau de bord"
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells("A1:H1")
ws1["A1"] = "ANALYSE DU MARCHE AIRBNB — DOUALA, CAMEROUN"
ws1["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
ws1["A1"].fill = fill(BLUE_DARK)
ws1["A1"].alignment = center()
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:H2")
ws1["A2"] = "Source : airbnb.com + donnees simulees | Base MySQL airbnb_douala | Juin 2026"
ws1["A2"].font = Font(name="Arial", size=9, italic=True, color="888888")
ws1["A2"].alignment = center()
ws1.row_dimensions[2].height = 16

# KPI boxes
ws1.row_dimensions[4].height = 18
ws1.row_dimensions[5].height = 32
ws1.row_dimensions[6].height = 18

kpis = [
    ("A", "C", "Total annonces", f"=COUNTA(Donnees_brutes!A:A)-1"),
    ("D", "F", "Annonces reelles", str(len(df_real))),
    ("G", "I", "Prix moyen / nuit (XAF)", f"=ROUND(AVERAGE(Donnees_brutes!E:E),0)"),
]

col_pairs = [("A","C"), ("D","F"), ("G","I")]
kpi_labels = ["Total annonces", "Annonces reelles Airbnb", "Prix moyen / nuit (XAF)"]
kpi_values = [
    f"=COUNTA(Donnees_brutes!A2:A5000)",
    str(len(df_real)),
    f"=ROUND(AVERAGEIF(Donnees_brutes!A2:A5000,\"<>\"&\"\",Donnees_brutes!E2:E5000),0)",
]

for (c1, c2), label, val in zip(col_pairs, kpi_labels, kpi_values):
    ws1.merge_cells(f"{c1}4:{c2}4")
    ws1[f"{c1}4"] = label
    ws1[f"{c1}4"].font = Font(name="Arial", size=9, color="888888")
    ws1[f"{c1}4"].fill = fill(BLUE_LIGHT)
    ws1[f"{c1}4"].alignment = center()

    ws1.merge_cells(f"{c1}5:{c2}5")
    ws1[f"{c1}5"] = val
    ws1[f"{c1}5"].font = Font(name="Arial", size=18, bold=True, color=BLUE_DARK)
    ws1[f"{c1}5"].fill = fill(BLUE_LIGHT)
    ws1[f"{c1}5"].alignment = center()

    ws1.merge_cells(f"{c1}6:{c2}6")
    ws1[f"{c1}6"].fill = fill(BLUE_MED)

# Section: Stats quartiers
ws1.row_dimensions[8].height = 20
ws1.merge_cells("A8:I8")
ws1["A8"] = "STATISTIQUES PAR QUARTIER"
ws1["A8"].font = Font(name="Arial", size=11, bold=True, color=WHITE)
ws1["A8"].fill = fill(BLUE_MED)
ws1["A8"].alignment = center()

hdr_q = ["Quartier", "Total annonces", "Prix min (XAF)", "Prix moyen (XAF)", "Prix max (XAF)", "Note moy.", "Annonces reelles"]
for col, h in enumerate(hdr_q, 1):
    c = ws1.cell(row=9, column=col, value=h)
    c.font = hdr_font(10)
    c.fill = fill(BLUE_DARK)
    c.alignment = center()
    c.border = border()

for i, row in df_quartiers.iterrows():
    r = 10 + i
    alt = i % 2 == 0
    vals = [row["neighborhood"], row["total_annonces"], row["prix_min"], row["prix_moyen"], row["prix_max"], row["note_moyenne"], row["annonces_reelles"]]
    for col, val in enumerate(vals, 1):
        c = ws1.cell(row=r, column=col, value=val)
        c.font = cell_font(10)
        c.fill = fill(GRAY_LIGHT if alt else WHITE)
        c.alignment = center()
        c.border = border()

# Totals row
tot_row = 10 + len(df_quartiers)
ws1.cell(row=tot_row, column=1, value="TOTAL").font = Font(name="Arial", size=10, bold=True)
ws1.cell(row=tot_row, column=1).fill = fill(BLUE_LIGHT)
ws1.cell(row=tot_row, column=2, value=f"=SUM(B10:B{tot_row-1})").font = Font(name="Arial", size=10, bold=True)
ws1.cell(row=tot_row, column=2).fill = fill(BLUE_LIGHT)
ws1.cell(row=tot_row, column=2).alignment = center()
for col in range(1, 8):
    ws1.cell(row=tot_row, column=col).border = border()

set_col_widths(ws1, {"A": 18, "B": 16, "C": 16, "D": 18, "E": 16, "F": 12, "G": 18})

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — ANNONCES REELLES
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Annonces reelles")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:H1")
ws2["A1"] = "ANNONCES REELLES AIRBNB — DOUALA (254 logements collectes)"
ws2["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws2["A1"].fill = fill(BLUE_DARK)
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 30

hdrs = ["ID", "Titre", "Quartier", "Type", "Prix/nuit (XAF)", "Note", "Avis", "Source"]
for col, h in enumerate(hdrs, 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.font = hdr_font(10)
    c.fill = fill(BLUE_MED)
    c.alignment = center()
    c.border = border()

for i, row in df_real.iterrows():
    r = 3 + i
    alt = i % 2 == 0
    vals = [row["id"], row["title"], row["neighborhood"], row["type"],
            row["price_per_night"], row["rating"], row["reviews"], row["source"]]
    for col, val in enumerate(vals, 1):
        c = ws2.cell(row=r, column=col, value=val)
        c.font = cell_font(9)
        c.fill = fill(GRAY_LIGHT if alt else WHITE)
        c.alignment = left() if col == 2 else center()
        c.border = border()

# Summary at bottom
last = 3 + len(df_real)
ws2.cell(row=last+1, column=1, value="Total annonces reelles").font = Font(name="Arial", size=10, bold=True, color=BLUE_DARK)
ws2.cell(row=last+1, column=5, value=f"=COUNTA(E3:E{last})").font = Font(name="Arial", size=10, bold=True)
ws2.cell(row=last+2, column=1, value="Prix moyen / nuit").font = Font(name="Arial", size=10, bold=True, color=BLUE_DARK)
ws2.cell(row=last+2, column=5, value=f"=ROUND(AVERAGE(E3:E{last}),0)").font = Font(name="Arial", size=10, bold=True)
ws2.cell(row=last+3, column=1, value="Note moyenne").font = Font(name="Arial", size=10, bold=True, color=BLUE_DARK)
ws2.cell(row=last+3, column=6, value=f"=ROUND(AVERAGE(F3:F{last}),2)").font = Font(name="Arial", size=10, bold=True)

set_col_widths(ws2, {"A": 12, "B": 40, "C": 16, "D": 22, "E": 18, "F": 10, "G": 10, "H": 12})

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — DONNEES BRUTES
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Donnees_brutes")
cols_export = ["id", "title", "neighborhood", "type", "price_per_night", "rating", "reviews", "occupancy_rate", "monthly_revenue", "superhost", "source"]
df_export = df_all[cols_export].copy()

hdrs3 = ["ID", "Titre", "Quartier", "Type", "Prix/nuit XAF", "Note", "Avis", "Taux occupation", "Revenu mensuel", "Superhost", "Source"]
for col, h in enumerate(hdrs3, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = hdr_font(10)
    c.fill = fill(BLUE_DARK)
    c.alignment = center()
    c.border = border()

for i, row in df_export.iterrows():
    r = 2 + i
    for col, val in enumerate(row, 1):
        c = ws3.cell(row=r, column=col, value=val)
        c.font = cell_font(9)
        c.alignment = left() if col == 2 else center()

set_col_widths(ws3, {"A": 12, "B": 40, "C": 16, "D": 22, "E": 16, "F": 10, "G": 8, "H": 16, "I": 16, "J": 12, "K": 12})

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — STATS TYPES
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Stats par type")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:D1")
ws4["A1"] = "REPARTITION PAR TYPE DE LOGEMENT"
ws4["A1"].font = Font(name="Arial", size=12, bold=True, color=WHITE)
ws4["A1"].fill = fill(BLUE_DARK)
ws4["A1"].alignment = center()
ws4.row_dimensions[1].height = 28

hdrs4 = ["Type de bien", "Nombre", "Prix moyen (XAF)", "Note moyenne"]
for col, h in enumerate(hdrs4, 1):
    c = ws4.cell(row=2, column=col, value=h)
    c.font = hdr_font(10)
    c.fill = fill(BLUE_MED)
    c.alignment = center()
    c.border = border()

for i, row in df_types.iterrows():
    r = 3 + i
    alt = i % 2 == 0
    vals = [row["type"], row["total"], row["prix_moyen"], row["note_moyenne"]]
    for col, val in enumerate(vals, 1):
        c = ws4.cell(row=r, column=col, value=val)
        c.font = cell_font(10)
        c.fill = fill(GRAY_LIGHT if alt else WHITE)
        c.alignment = center()
        c.border = border()

tot4 = 3 + len(df_types)
ws4.cell(row=tot4, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
ws4.cell(row=tot4, column=1).fill = fill(BLUE_LIGHT)
ws4.cell(row=tot4, column=2, value=f"=SUM(B3:B{tot4-1})").font = Font(name="Arial", bold=True, size=10)
ws4.cell(row=tot4, column=2).fill = fill(BLUE_LIGHT)
ws4.cell(row=tot4, column=2).alignment = center()
for col in range(1, 5):
    ws4.cell(row=tot4, column=col).border = border()

set_col_widths(ws4, {"A": 25, "B": 14, "C": 20, "D": 16})

# Add bar chart on sheet 4
chart = BarChart()
chart.type = "col"
chart.title = "Prix moyen par type de logement (XAF)"
chart.y_axis.title = "Prix (XAF)"
chart.x_axis.title = "Type"
chart.style = 10
data = Reference(ws4, min_col=3, min_row=2, max_row=2+len(df_types))
cats = Reference(ws4, min_col=1, min_row=3, max_row=2+len(df_types))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
chart.width = 18
chart.height = 12
ws4.add_chart(chart, "F2")

wb.save(str(OUTPUT))
print("Excel file saved: " + str(OUTPUT))
print("Sheets: Tableau de bord, Annonces reelles, Donnees_brutes, Stats par type")
